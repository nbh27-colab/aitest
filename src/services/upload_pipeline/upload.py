from typing import List, Optional
from pydantic import BaseModel
from openpyxl import load_workbook
from collections import defaultdict
import os
from sqlalchemy.future import select

from config.settings import MinIOSettings
from utils.s3_client import get_s3
from src.data.database.crud.dbs_manager import AsyncDatabaseManager
from src.models import Project, CaseFile, CaseSheet, Step
from src.models.test_case import TestCase as TestCaseORM

class TestStep(BaseModel):
    step: str
    expected_result: str
    comment: Optional[str] = None

class TestCase(BaseModel):
    test_case_title: str
    steps: List[TestStep]
    sheet_name: Optional[str] = None


class UploadService:
    @staticmethod
    async def insert_test_cases(file_path: str, project_id: int):
        test_cases: List[TestCase] = UploadService.parse_test_cases(file_path)
        filename = os.path.basename(file_path)

        async with AsyncDatabaseManager().connect_session_async() as session:
            project = await session.scalar(
                select(Project).where(Project.project_id == project_id)
            )
            if not project:
                raise ValueError(f"Project with id {project_id} does not exist.")
            
            case_file = CaseFile(
                project_id=project_id,
                name=filename,
                file_path=file_path
            )
            session.add(case_file)
            await session.flush()

            sheet_map = {}

            for tc in test_cases:
                if tc.sheet_name not in sheet_map:
                    case_sheet = CaseSheet(
                        case_file_id=case_file.case_file_id,
                        name=tc.sheet_name
                    )
                    session.add(case_sheet)
                    await session.flush()
                    sheet_map[tc.sheet_name] = case_sheet.case_sheet_id

                sheet_id = sheet_map[tc.sheet_name]

                existing_tc = await session.scalar(
                    select(TestCaseORM).where(
                        TestCaseORM.case_sheet_id == sheet_id,
                        TestCaseORM.title == tc.test_case_title
                    )
                )
                if existing_tc:
                    test_case_orm = existing_tc
                else:
                    test_case_orm = TestCaseORM(
                        case_sheet_id=sheet_id,
                        title=tc.test_case_title
                    )
                    session.add(test_case_orm)
                    await session.flush()

                for order, step in enumerate(tc.steps, start=1):
                    existing_step = await session.scalar(
                        select(Step).where(
                            Step.test_case_id == test_case_orm.test_case_id,
                            Step.step_order == order,
                            Step.action == step.step,
                        )
                    )
                    if existing_step:
                        continue
                    step = Step(
                        test_case_id=test_case_orm.test_case_id,
                        project_id=project_id,
                        step_order=order,
                        action=step.step,
                        expected_result=step.expected_result,
                        comment=step.comment
                    )
                    session.add(step)
            await session.commit()
            print(f"[INFO] Inserted test cases from {filename} into project {project_id}")
            return case_file.case_file_id


    @staticmethod
    def parse_test_cases(file_path: str) -> List[TestCase]:
        s3 = get_s3()
        settings = MinIOSettings()
        bucket_name = settings.BUCKET_NAME
        remote_path = file_path.lstrip("/")

        print(f"[DEBUG] Downloading file from S3: bucket={bucket_name}, path={remote_path}")

        with s3.download_file(
            bucket_name=bucket_name,
            remote_file_path=remote_path
        ) as local_file:
            ext = local_file.suffix.lower()
            if ext == ".xlsx":
                return UploadService.parse_xlsx(local_file)
            else:
                raise ValueError(f"Unsupported file extension: {ext}")
            

    @staticmethod
    def parse_xlsx(file_path: str) -> List[TestCase]:
        wb = load_workbook(file_path, data_only=True)
        all_cases: List[TestCase] = []

        for sheet in wb.sheetnames:
            ws = wb[sheet]
            headers = [cell.value for cell in ws[1]]
            last_case_title = None
            grouped_data = defaultdict(list)

            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                row_dict = {
                    str(k).strip(): v for k, v in row_dict.items() if k is not None
                }

                if row_dict.get("Case Title"):
                    last_case_title = row_dict["Case Title"]

                grouped_data[last_case_title].append(
                    TestStep(
                        step=str(row_dict.get("Steps") or ""),
                        expected_result=str(row_dict.get("Expected Result") or ""),
                        comment=str(row_dict.get("Comments") or "")
                    )
                )

            for title, steps in grouped_data.items():
                if not title:
                    continue
                all_cases.append(
                    TestCase(
                        test_case_title=title,
                        steps=steps,
                        sheet_name=sheet
                    )
                )
        return all_cases
    
